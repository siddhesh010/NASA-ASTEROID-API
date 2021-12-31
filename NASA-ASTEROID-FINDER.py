#Copyright <2021> <SIDDHESH VIKAS NAIK>
#Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the "Software"), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:
#The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.
#THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

#This software is based on NASA API-Asteroids - NeoWs
#The above mentioned API is maintained by SpaceRocks https://github.com/SpaceRocks/
#For more information visit:- https://api.nasa.gov/

from PyQt5 import QtCore, QtGui, QtWidgets
import requests
import json
from openpyxl import Workbook
import os



class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 400)
        MainWindow.setMinimumSize(QtCore.QSize(800, 400))
        MainWindow.setMaximumSize(QtCore.QSize(800, 400))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(100)
        MainWindow.setFont(font)
        MainWindow.setAnimated(True)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)     # TOP LABEL
        self.label.setGeometry(QtCore.QRect(0, 0, 800, 20))
        self.label.setMinimumSize(QtCore.QSize(800, 20))
        self.label.setMaximumSize(QtCore.QSize(800, 20))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.label.setFont(font)
        self.label.setMouseTracking(True)
        self.label.setTabletTracking(True)
        self.label.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        

        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)          # API KEY TEXT BOX
        self.textEdit.setGeometry(QtCore.QRect(5, 20, 790, 50))
        self.textEdit.setObjectName("textEdit")

        
        self.dateEdit = QtWidgets.QDateEdit(self.centralwidget)           # START DATE WIDGET
        self.dateEdit.setGeometry(QtCore.QRect(5, 80, 140, 60))
        self.dateEdit.setObjectName("dateEdit")
        

        self.dateEdit_2 = QtWidgets.QDateEdit(self.centralwidget)          #END DATE WIDGET
        self.dateEdit_2.setGeometry(QtCore.QRect(160, 80, 140, 60))
        self.dateEdit_2.setObjectName("dateEdit_2")
        

        self.label_2 = QtWidgets.QLabel(self.centralwidget)                # START DATE LABEL
        self.label_2.setGeometry(QtCore.QRect(26, 140, 51, 16))
        self.label_2.setObjectName("label_2")
        self.label_2.setFixedSize(100,16)

        
        self.label_3 = QtWidgets.QLabel(self.centralwidget)                 # END DATE LABEL
        self.label_3.setGeometry(QtCore.QRect(190, 140, 61, 16))
        self.label_3.setObjectName("label_3")
        self.label_3.setFixedSize(100,16)

        
        self.textEdit_2 = QtWidgets.QTextEdit(self.centralwidget)           # FILE NAME TEXT BOX
        self.textEdit_2.setGeometry(QtCore.QRect(330, 80, 300, 40))
        self.textEdit_2.setObjectName("textEdit_2")

        
        self.textEdit_3 = QtWidgets.QTextEdit(self.centralwidget)           # REQUEST STATUS TEXT BOX
        self.textEdit_3.setGeometry(QtCore.QRect(5, 200, 800, 200))
        self.textEdit_3.setObjectName("textEdit_3")

        
        self.label_4 = QtWidgets.QLabel(self.centralwidget)                # FILE NAME LABEL
        self.label_4.setGeometry(QtCore.QRect(400, 130, 110, 40))
        self.label_4.setObjectName("label_4")
        self.label_4.setFixedSize(150,16)

        
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)        #  GET DATA BUTTON
        self.pushButton.setGeometry(QtCore.QRect(650, 80, 100, 40))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.get_data)

        
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "NASA-Asteroids-NeoWs(Near Earth Object Web Service)"))
        self.label.setText(_translate("MainWindow", "NASA API KEY"))
        self.label_2.setText(_translate("MainWindow", "Start Date"))
        self.label_3.setText(_translate("MainWindow", "End Date"))
        self.label_4.setText(_translate("MainWindow", "  Save as Filename"))
        self.textEdit_3.setText(_translate("MainWindow", "REQUEST STATUS:-\n\n"))
        self.pushButton.setText(_translate("MainWindow", "Get Data"))

    def get_data(self):
        api_key=self.textEdit.toPlainText()
        start_date=str(self.dateEdit.text()[6:10])+str(self.dateEdit.text()[2:6])+str(self.dateEdit.text()[0:2])    #Formatting the string according to API Format
        end_date=str(self.dateEdit_2.text()[6:10])+str(self.dateEdit_2.text()[2:6])+str(self.dateEdit_2.text()[0:2])  
        url="https://api.nasa.gov/neo/rest/v1/feed?start_date={}&end_date={}&api_key={}".format(start_date,end_date,api_key)  #making the url according to format
        
        if len(str(self.textEdit_2.toPlainText()))>0:
            filename=str(self.textEdit_2.toPlainText())+".xlsx"
        else:
            filename="ASTEROID_DATA_NASA.xlsx"


        
        data=requests.get(url)   # getting data from NASA API

        if data.status_code==200:
            self.textEdit_3.setText("REQUEST STATUS:-\n\nREQUEST SUCCESSFUL!!")
            
            json_data=json.loads(data.text)    

            near_earth=json_data["near_earth_objects"]
            dates=[]
            for i in near_earth.keys():
                dates.append(i)
            wb = Workbook()
            asteroids=[]

            for date in dates:
                for j in near_earth[date]:
                    asteroids.append(j)
            wb = Workbook()
            ws=wb.active

            self.textEdit_3.setText("REQUEST STATUS:-\n\nCleaning data...")
           

            ws["A1"]="link"
            ws["B1"]="Id"
            ws["C1"]="Name"
            ws["D1"]="Nasa_Jpl_Url"
            ws["E1"]="Absolute_Magnitude_h"
            ws["F1"]="Estimated_Diameter_Min(Km)"
            ws["G1"]="Estimated_Diameter_Max(Km)"
            ws["H1"]="Close_Approach_Date"
            ws["I1"]="Relative_Velocity(in Km)"
            ws["J1"]="Miss_Distance(in Km)"
            ws["K1"]="Orbiting_Body"

            self.textEdit_3.setText("Request Status:-\nSetting up Excel Sheet...")
            
            for i in range(0,len(asteroids)):
                ws["A"+str(i+2)]=asteroids[i]["links"]["self"]
                ws["B"+str(i+2)]=float(asteroids[i]["id"])
                ws["C"+str(i+2)]=asteroids[i]["name"]
                ws["D"+str(i+2)]=asteroids[i]["nasa_jpl_url"]
                ws["E"+str(i+2)]=float(asteroids[i]["absolute_magnitude_h"])
                ws["F"+str(i+2)]=float(asteroids[i]["estimated_diameter"]["kilometers"]["estimated_diameter_min"])
                ws["G"+str(i+2)]=float(asteroids[i]["estimated_diameter"]["kilometers"]["estimated_diameter_max"])
                ws["H"+str(i+2)]=asteroids[i]["close_approach_data"][0]["close_approach_date_full"]
                ws["I"+str(i+2)]=float(asteroids[i]["close_approach_data"][0]["relative_velocity"]["kilometers_per_second"])
                ws["J"+str(i+2)]=float(asteroids[i]["close_approach_data"][0]["miss_distance"]["kilometers"])
                ws["K"+str(i+2)]=asteroids[i]["close_approach_data"][0]["orbiting_body"]
            
            wb.save(filename)
            path="{} in {}".format(filename,str(os.getcwd()))
            self.textEdit_3.setText("REQUEST STATUS:-\n\nDONE\nFile saved as "+path)
           
        else:
            self.textEdit_3.setText("REQUEST STATUS:-\n\nERROR CODE {} \n{}".format(str(data.status_code),data.content))
            #print(url)
            
            


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
